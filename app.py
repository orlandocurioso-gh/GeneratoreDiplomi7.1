from flask import Flask, request, send_file, render_template, redirect, url_for
from weasyprint import HTML, CSS
from pypdf import PdfWriter
import io
import csv
import zipfile
import tempfile
import shutil
import os
from datetime import datetime
import threading
import uuid
import openpyxl
from openpyxl import Workbook, load_workbook
import pathlib

# --- CONFIGURAZIONE PERCORSI RELATIVI ---
# Rileva la cartella dove si trova app.py
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Crea i percorsi relativi alla cartella del progetto
PATH_ARCHIVIO_1 = os.path.join(BASE_DIR, "Archivio_Locale")
PATH_ARCHIVIO_2 = os.path.join(BASE_DIR, "Archivio_Franco")
PATH_EXCEL_REGISTRO = os.path.join(BASE_DIR, "Registri")

# Aggiungi questo percorso
PATH_STAMPA = os.path.join(BASE_DIR, "Da_Stampare")

# Assicurati che esista all'avvio
for p in [PATH_ARCHIVIO_1, PATH_ARCHIVIO_2, PATH_EXCEL_REGISTRO, PATH_STAMPA]:
    os.makedirs(p, exist_ok=True)

# Assicurati che le cartelle esistano all'avvio
for p in [PATH_ARCHIVIO_1, PATH_ARCHIVIO_2, PATH_EXCEL_REGISTRO]:
    os.makedirs(p, exist_ok=True)

# --- FUNZIONI DI UTILITÀ ---
def format_place_name(place_str):
    if not place_str: return ""
    # Preposizioni e articoli che devono restare minuscoli
    PREPOSITIONS = {
        'di', 'del', 'della', 'degli', 'dei', 'de', 'da', 'dal', 
        'dalla', 'dai', 'dagli', 'su', 'sul', 'sulla', 'sui', 
        'sugli', 'a', 'al', 'alla', 'ai', 'agli', 'in', 'nel', 
        'nella', 'nei', 'negli', 'per', 'con', 'e', 'il', 'lo', 'la', 'gli', 'le', 
        "d'", "l'","de'",'val','meno',"all'"
    }
    # Articoli con apostrofo
    APOSTROPHES = {"d'", "l'", "all'", "dell'", "nell'"}
    
    words = place_str.lower().split()
    final = []
    
    for i, word in enumerate(words):
        # Se è una preposizione e non è la prima parola, resta minuscola
        if word in PREPOSITIONS and i > 0:
            final.append(word.lower())
        # Gestione apostrofi (es: l'aquila -> L'Aquila)
        elif "'" in word:
            parts = word.split("'")
            prefix = parts[0] + "'"
            suffix = parts[1]
            if prefix in APOSTROPHES and i > 0:
                final.append(prefix.lower() + suffix.capitalize())
            else:
                final.append(prefix.capitalize() + suffix.capitalize())
        # Altrimenti capitalizza normalmente
        else:
            final.append(word.capitalize())
            
    return ' '.join(final)

def format_name_with_exceptions(name_str):
    if not name_str: return ""
    words = name_str.split()
    return ' '.join([w[1:].lower() if w.startswith('%') else w.lower().capitalize() for w in words])

app = Flask(__name__, static_folder='static')
temp_pdf_batches = {}
CLEANUP_DELAY_SECONDS = 3600 

def cleanup_batch_data(batch_id):
    batch_info = temp_pdf_batches.pop(batch_id, None)
    if batch_info:
        try:
            shutil.rmtree(batch_info['temp_dir'])
        except Exception as e:
            print(f"Errore pulizia: {e}")

@app.route('/', methods=['GET'])
def homepage():
    return render_template('upload.html')

@app.route('/upload-data', methods=['POST'])
def upload_data():
    # 1. Recupero la facoltà dal menu a discesa (campo obbligatorio)
    facolta_selezionata = request.form.get('facolta_selezionata')
    
    if 'data_file' not in request.files or not facolta_selezionata:
        return 'Dati mancanti: seleziona la facoltà e carica il file.', 400

    file = request.files['data_file']
    if file.filename == '':
        return 'Nessun file selezionato', 400

    file_content = file.stream.read().decode('utf-8')
    students_data = parse_diploma_data(file_content)

    if not students_data:
        return 'File dati non valido o vuoto.', 400

    batch_id = str(uuid.uuid4())
    current_batch_temp_dir = tempfile.mkdtemp()
    generated_pdf_filenames = []
    log_entries = []
    nome_cartella = datetime.now().strftime('%Y-%m-%d')

    # --- CICLO GENERAZIONE PDF ---
    for i, student in enumerate(students_data):
        student_data_for_template = {k.lower(): v for k, v in student.items()}


    # --- FORMATTAZIONE NOMI E LUOGHI ---
        
    # Gestione a capo e nomi (Logica originale)
        student_data_for_template['corsolau'] = student_data_for_template.get('corsolau', '').replace('|', '<br>')
        student_data_for_template['nom_cog'] = student_data_for_template.get('nom_cog', '').replace('|', '<br>')
        
        # --- LOGICA LUOGO DI NASCITA (Ripristinata) ---
        luogo_nascita_completo = student_data_for_template.get('luogonas', '').strip()
        stato_nascita = student_data_for_template.get('statnas', '').strip()
        provincia_nascita = student_data_for_template.get('provnas', '').strip()
        
        if provincia_nascita:
            luogo_nascita_completo += f" {provincia_nascita}"
            
        if stato_nascita:
            luogo_nascita_completo += f" {stato_nascita}"
        
        # Sovrascriviamo per il template
        student_data_for_template['luogonas'] = luogo_nascita_completo
        # -----------------------------------------------



        modulo_value = student_data_for_template.get('modulo', '').strip()
        templates = {
            'forml01v7': 'diploma_forml01v7.html',
            'forml01v7tuscia': 'diploma_forml01v7tuscia.html',
            'forml1v7': 'diploma_forml1v7.html',
            'forml2v7': 'diploma_forml2v7.html',
            'forml3v7': 'diploma_forml3v7.html',
            'forml4v7': 'diploma_forml4v7.html',
            'forml23v7': 'diploma_forml23v7.html',
            'forml27v7': 'diploma_forml27v7.html',
            'forml28v7': 'diploma_forml28v7.html',
            'forml28v7A': 'diploma_forml28v7A.html',
            'forml29v7': 'diploma_forml29v7.html',
            'memoriastudi': 'diploma_memoriastudi.html',
            'memorialaureamag': 'diploma_memorialaureamag.html',
            'memorialaureatri': 'diploma_memorialaureatri.html'
        }

        template_filename = templates.get(modulo_value)
        if not template_filename:
            log_entries.append(f"SKIP: Modulo '{modulo_value}' non trovato per {student_data_for_template.get('nom_cog')}")
            continue

        student_data_for_template['lode'] = student_data_for_template.get('lode', '').upper().strip()
        student_data_for_template['testo_footer_fisso'] = "Imposta di bollo assolta in modo virtuale. Autorizzazione Intendenza di Finanza di Roma n.9120/88"


        # Gestione immagini firme/loghi
        for k in ['firmar', 'firmap', 'firmad', 'firma4', 'firma5', 'firma6', 'logo1', 'logo2', 'logo3']:
            val = student_data_for_template.get(k)
            if val and not val.endswith('.png'):
                student_data_for_template[k] = f"{val}.png"

        try:
            # Generazione Diploma
            rendered_html = render_template(template_filename, **student_data_for_template)
            pdf_bytes = HTML(string=rendered_html, base_url=request.url_root).write_pdf()
            clean_name = student_data_for_template.get('nom_cog', 'studente').replace(' ', '_').replace('<br>', '_')
            pdf_name = f'diploma_{clean_name}_{modulo_value}.pdf'
            
            with open(os.path.join(current_batch_temp_dir, pdf_name), 'wb') as f:
                f.write(pdf_bytes)
            generated_pdf_filenames.append(pdf_name)

            # Generazione Camicia
            camicia_data = {
                'corso_laurea': student_data_for_template.get('corsolau', ''),
                'nome_studente': student_data_for_template.get('nom_cog', ''),
                'luogo_nascita': luogo_nascita_completo,
                'provincia_nascita': provincia_nascita,
                'data_nascita': student_data_for_template.get('datanas', ''),
                'numero_protocollo': student_data_for_template.get('protocol', ''),
                'numero_diploma': student_data_for_template.get('npergamena', ''),
                'genere_nato_nata': student_data_for_template.get('sesso', 'nato/a').strip(),
                'firmad': student_data_for_template.get('firmad', ''),
                'firmar': student_data_for_template.get('firmar', ''),
                'firmap': student_data_for_template.get('firmap', '')
            }
            c_html = render_template('camicia_template.html', **camicia_data)
            c_pdf_bytes = HTML(string=c_html, base_url=request.url_root).write_pdf()
            c_pdf_name = f'camicia_{clean_name}.pdf'
            
            with open(os.path.join(current_batch_temp_dir, c_pdf_name), 'wb') as f:
                f.write(c_pdf_bytes)
            generated_pdf_filenames.append(c_pdf_name)
            
            log_entries.append(f"OK: {student_data_for_template.get('nom_cog')}")
        except Exception as e:
            log_entries.append(f"ERRORE {student_data_for_template.get('nom_cog')}: {e}")

    # --- OPERAZIONI POST-GENERAZIONE ---
    # Merge Diplomi
    diploma_files = [f for f in generated_pdf_filenames if f.startswith('diploma_')]
    if diploma_files:
        merger = PdfWriter()
        comb_name = f'tutti_i_diplomi_{nome_cartella}.pdf'
        for f in diploma_files:
            merger.append(os.path.join(current_batch_temp_dir, f))
        merger.write(os.path.join(current_batch_temp_dir, comb_name))
        merger.close()
        generated_pdf_filenames.append(comb_name)
        
    camicia_files = [f for f in generated_pdf_filenames if f.startswith('camicia_')]
    if camicia_files:
        merger_c = PdfWriter()
        comb_c_name = f'tutte_le_camicie_{nome_cartella}.pdf'
        for f in camicia_files:
            merger_c.append(os.path.join(current_batch_temp_dir, f))
        merger_c.write(os.path.join(current_batch_temp_dir, comb_c_name))
        merger_c.close()
        generated_pdf_filenames.append(comb_c_name)

    # Definizione log_content (FIX UnboundLocalError)
    log_content = '\n'.join(log_entries)
    log_file_path = os.path.join(current_batch_temp_dir, 'log_creazione_diplomi.txt')
    with open(log_file_path, 'w', encoding='utf-8') as f:
        f.write(log_content)

    # Preparazione Metadati per Archivio
    primo_studente = students_data[0] if students_data else {}
    
    # 1. Protocollo pulito (es. 16828/1 -> 16828)
    prot_raw = primo_studente.get('PROTOCOL', '').strip()
    protocollo_clean = prot_raw.split('/')[0] if '/' in prot_raw else prot_raw
    
    # 2. Tipologia (LM- in CLASSE -> LaureaMagistrale)
    classe_val = primo_studente.get('CLASSE', '')
    tipologia = "LaureaMagistrale" if "LM-" in classe_val else "LaureaTriennale"
    
    # 3. Anno Laurea
    anno_lau = primo_studente.get('DATALAUR', datetime.now().strftime('%Y')).strip().replace('/', '-')

    # Salvataggio Batch
    temp_pdf_batches[batch_id] = {
        'temp_dir': current_batch_temp_dir,
        'filenames': generated_pdf_filenames,
        'log_content': log_content,
        'log_file_path': log_file_path,
        'original_folder_name': nome_cartella,
        'archived': False,
        'metadata': {
            'protocollo': protocollo_clean,
            'tipologia': tipologia,
            'facolta': facolta_selezionata.replace(' ', '_'),
            'anno_laurea': anno_lau,
            'nomi_persone': [s.get('NOM_COG', 'N/A') for s in students_data],
            'totale': len(students_data),
            'student_list': students_data
        }
    }
    
    threading.Timer(CLEANUP_DELAY_SECONDS, cleanup_batch_data, args=[batch_id]).start()
    return redirect(url_for('preview_pdfs', batch_id=batch_id))

@app.route('/archive/<batch_id>', methods=['POST'])
def archive_batch(batch_id):
    batch_info = temp_pdf_batches.get(batch_id)
    if not batch_info or batch_info.get('archived'):
        return "Batch non trovato o già archiviato.", 404

    meta = batch_info['metadata']
    now = datetime.now()
    timestamp = now.strftime('%d%m%Y_%H%M')
    
    # Nome cartella e ZIP
    folder_name = f"{meta['tipologia']}_{meta['totale']}_{meta['facolta']}_{meta['anno_laurea']}_{timestamp}"
    zip_name = f"{folder_name}.zip"
    temp_zip_path = os.path.join(batch_info['temp_dir'], zip_name)

    try:
        # Crea ZIP
        with zipfile.ZipFile(temp_zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for f in batch_info['filenames']:
                if f.startswith('diploma_'):
                    zf.write(os.path.join(batch_info['temp_dir'], f), arcname=os.path.join(folder_name, f))
            # Aggiunge log nominativi
            log_arc_path = os.path.join(batch_info['temp_dir'], "log_archivio.txt")
            with open(log_arc_path, 'w', encoding='utf-8') as la:
                la.write(f"REGISTRO {meta['tipologia']} - {now}\n" + "\n".join(meta['nomi_persone']))
            zf.write(log_arc_path, arcname=os.path.join(folder_name, "lista_nomi.txt"))

        # Copia nei server
        shutil.copy2(temp_zip_path, os.path.join(PATH_ARCHIVIO_1, zip_name))
        shutil.copy2(temp_zip_path, os.path.join(PATH_ARCHIVIO_2, zip_name))

        # Excel
        excel_path = os.path.join(PATH_EXCEL_REGISTRO, f"Pergamene_{now.year}.xlsx")
        new_row = [meta['protocollo'], meta['tipologia'], meta['totale'], meta['facolta'], meta['anno_laurea'], now.strftime('%d/%m/%Y %H:%M')]
        
        if not os.path.exists(excel_path):
            wb = Workbook()
            ws = wb.active
            ws.append(["Protocollo", "Tipologia", "Totale PDF", "Facoltà", "Anno Laurea", "Data Stampa"])
        else:
            wb = load_workbook(excel_path)
            ws = wb.active
        
        ws.append(new_row)
        wb.save(excel_path)
        
        batch_info['archived'] = True
        return f"Archiviazione completata. Protocollo: {meta['protocollo']}", 200
    except Exception as e:
        return f"Errore archivio: {str(e)}", 500

@app.route('/preview/<batch_id>')
def preview_pdfs(batch_id):
    batch_info = temp_pdf_batches.get(batch_id)
    if not batch_info:
        return "Anteprima non trovata o scaduta.", 404

    pdf_list_for_template = []
    # Prepara la lista di PDF per il template, includendo SOLO i diplomi
    for filename in batch_info['filenames']:
        # Aggiungi questa condizione per filtrare solo i PDF dei diplomi
        if filename.startswith('diploma_'): 
            pdf_list_for_template.append({
                'name': filename,
                'url': url_for('get_single_pdf', batch_id=batch_id, filename=filename)
            })

    return render_template('preview.html',
                            pdf_list=pdf_list_for_template,
                            download_url=url_for('download_zip_for_preview', batch_id=batch_id),
                            log_url=url_for('get_log_for_preview', batch_id=batch_id),
                            cleanup_delay_minutes=CLEANUP_DELAY_SECONDS / 60)

@app.route('/preview/pdf/<batch_id>/<filename>')
def get_single_pdf(batch_id, filename):
    batch_info = temp_pdf_batches.get(batch_id)
    if not batch_info:
        return "File non trovato.", 404
    
    if filename not in batch_info['filenames']:
        return "File non autorizzato o non trovato nel batch.", 403

    return send_file(os.path.join(batch_info['temp_dir'], filename), mimetype='application/pdf')

@app.route('/preview/log/<batch_id>')
def get_log_for_preview(batch_id):
    batch_info = temp_pdf_batches.get(batch_id)
    if not batch_info:
        return "Log non trovato o scaduto.", 404
    
    return send_file(batch_info['log_file_path'], 
                    mimetype='text/plain', 
                    as_attachment=True, 
                    download_name='log_creazione_diplomi.txt')


@app.route('/download_zip/<batch_id>')
def download_zip_for_preview(batch_id):
    batch_info = temp_pdf_batches.get(batch_id)
    if not batch_info:
        return "Download non trovato o scaduto.", 404

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'a', zipfile.ZIP_DEFLATED, False) as zf:
        for filename in batch_info['filenames']:
            file_path = os.path.join(batch_info['temp_dir'], filename)
            
            if filename.startswith('diploma_'):
                subfolder = 'pergamene'
            elif filename.startswith('camicia_'):
                subfolder = 'camicie'
            elif filename.startswith('tutti_i_diplomi_'):
                subfolder = 'combinato'
            else:
                subfolder = 'altri' 

            arcname = os.path.join(batch_info['original_folder_name'], subfolder, filename)
            zf.write(file_path, arcname=arcname)
        
        log_filename_in_zip = os.path.join(batch_info['original_folder_name'], 'log_creazione_documenti.txt')
        zf.write(batch_info['log_file_path'], arcname=log_filename_in_zip)
    
    zip_buffer.seek(0)
    
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'documenti_{batch_info["original_folder_name"]}.zip'
    )

def parse_diploma_data(file_content):
    lines = file_content.splitlines()
    if len(lines) < 5: return []
    header_line = lines[3]
    data_lines = lines[4:]
    reader = csv.reader(io.StringIO(header_line + '\n' + '\n'.join(data_lines)), delimiter='^')
    try:
        headers = next(reader)
        return [{headers[i].strip(): row[i].strip() for i in range(len(headers))} for row in reader if len(row) == len(headers)]
    except: return []

@app.route('/print-files/<batch_id>', methods=['POST'])
def print_batch(batch_id):
    batch_info = temp_pdf_batches.get(batch_id)
    if not batch_info:
        return "Batch non trovato.", 404

    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        folder_name = f"Stampa_{batch_info['metadata']['facolta']}_{timestamp}"
        dest_path = os.path.join(PATH_STAMPA, folder_name)
        os.makedirs(dest_path, exist_ok=True)

        # Copia solo i file cumulativi (quelli che iniziano con 'tutti_' o 'tutte_')
        for filename in batch_info['filenames']:
            if filename.startswith('tutti_i_diplomi_') or filename.startswith('tutte_le_camicie_'):
                src = os.path.join(batch_info['temp_dir'], filename)
                shutil.copy2(src, os.path.join(dest_path, filename))

        # Generazione del file Elenco.txt (rimane uguale)
        elenco_path = os.path.join(dest_path, "Elenco.txt")
        with open(elenco_path, 'w', encoding='utf-8') as f:
            f.write(f"ELENCO STAMPA - {batch_info['metadata']['facolta']} - {timestamp}\n")
            f.write("-" * 80 + "\n")
            for s in batch_info['metadata']['student_list']:
                matricola = s.get('MATRI', 'N/D')
                nome = s.get('NOM_COG', 'N/D').replace('|', ' ')
                protocollo = s.get('PROTOCOL', 'N/D')
                corso = s.get('CORSOLAU', 'N/D').replace('|', ' ')
                f.write(f"MATR: {matricola} | NOM: {nome} | PROT: {protocollo} | CORSO: {corso}\n")

        return f"Cartella creata con file unici ed elenco: {folder_name}", 200
    except Exception as e:
        return f"Errore stampa: {str(e)}", 500

if __name__ == '__main__':
    app.run(debug=True)